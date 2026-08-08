"""
NL2SQL 模块：Excel 结构化数据 → SQLite → 自然语言 → LLM 生成 SQL → 精确查询
替代原有的 JSON exact-match 方案，面向真实数据体量（百行~万行级）设计。
"""
import sqlite3
import json
import re
import logging
from pathlib import Path
from typing import List, Dict, Optional, Tuple

import config

logger = logging.getLogger(__name__)

# SQLite 数据库路径（与 app.db 分开，专用于考试数据）
EXAM_DB_PATH = config.DATA_DIR / "exam.db"

# === SQL 安全白名单 ===
# 只允许 SELECT，禁止任何修改/删除/DDL 操作
FORBIDDEN_SQL_KEYWORDS = [
    "INSERT", "UPDATE", "DELETE", "DROP", "ALTER", "CREATE",
    "TRUNCATE", "REPLACE", "ATTACH", "DETACH", "PRAGMA",
    "EXEC", "EXECUTE", "--", "/*", "*/", "xp_", "sp_",
]

# 最大返回行数
MAX_RESULT_ROWS = 50


class ExamDatabase:
    """考研数据 SQLite 数据库管理器。

    每个 Excel sheet → 一张 SQLite 表。
    自动从 headers 推断列结构，支持增量导入去重。
    """

    def __init__(self, db_path: Path = None):
        self.db_path = db_path or EXAM_DB_PATH
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        self._init_meta()
        logger.info("ExamDatabase initialized at %s", self.db_path)

    def _get_conn(self) -> sqlite3.Connection:
        conn = sqlite3.connect(str(self.db_path))
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")
        return conn

    def _init_meta(self):
        """初始化元数据表——记录哪些 Excel sheet 已导入。"""
        conn = self._get_conn()
        conn.execute("""
            CREATE TABLE IF NOT EXISTS _exam_meta (
                table_name TEXT PRIMARY KEY,
                source_file TEXT NOT NULL,
                sheet_name TEXT NOT NULL,
                column_count INTEGER,
                row_count INTEGER,
                file_hash TEXT,
                indexed_at TEXT DEFAULT (datetime('now'))
            )
        """)
        conn.commit()
        conn.close()

    # ===== 导入 =====

    def ingest_excel(
        self,
        file_path: Path,
        force: bool = False,
    ) -> dict:
        """导入一个 Excel 文件的所有 sheet 到 SQLite。

        Args:
            file_path: Excel 文件路径
            force: 是否强制重新导入（忽略文件哈希）

        Returns:
            dict: {status, tables_created, total_rows, message}
        """
        if not file_path.exists():
            return {"status": "error", "message": f"文件不存在: {file_path}"}

        file_hash = self._file_hash(file_path)

        # 检查是否已导入
        if not force:
            existing = self._get_existing_hash(file_path.name)
            if existing and existing == file_hash:
                logger.info("Exam DB: skipping unchanged file %s", file_path.name)
                return {
                    "status": "skipped",
                    "message": f"文件未变更: {file_path.name}",
                    "hash": file_hash[:8],
                }

        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(file_path), data_only=True)
        except Exception as e:
            logger.exception("Failed to open Excel: %s", file_path)
            return {"status": "error", "message": f"Excel 打开失败: {e}"}

        tables_created = []
        total_rows = 0
        conn = self._get_conn()

        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows or len(rows) < 2:
                    continue

                # 清洗表名：去特殊字符、截断
                table_name = self._sanitize_table_name(sheet_name)

                # 清洗列名
                headers = [
                    self._sanitize_column_name(str(h) if h else f"col_{i}")
                    for i, h in enumerate(rows[0])
                ]

                # 建表（列宽 TEXT——数据清洗由 LLM 负责）
                col_defs = ", ".join(f'"{h}" TEXT' for h in headers)
                conn.execute(f'CREATE TABLE IF NOT EXISTS "{table_name}" (id INTEGER PRIMARY KEY AUTOINCREMENT, {col_defs})')

                # 清空旧数据（如果是 force 重新导入）
                if force:
                    conn.execute(f'DELETE FROM "{table_name}"')

                # 插入数据行
                placeholders = ", ".join("?" for _ in headers)
                col_names = ", ".join(f'"{h}"' for h in headers)
                insert_sql = f'INSERT INTO "{table_name}" ({col_names}) VALUES ({placeholders})'

                inserted = 0
                for row in rows[1:]:
                    if any(cell is not None for cell in row):
                        values = [str(cell) if cell is not None else "" for cell in row]
                        # 补齐长度
                        while len(values) < len(headers):
                            values.append("")
                        values = values[:len(headers)]
                        try:
                            conn.execute(insert_sql, values)
                            inserted += 1
                        except Exception as e:
                            logger.warning("Row insert failed in %s: %s", table_name, e)

                conn.commit()

                # 更新元数据
                conn.execute(
                    """INSERT OR REPLACE INTO _exam_meta
                       (table_name, source_file, sheet_name, column_count, row_count, file_hash)
                       VALUES (?, ?, ?, ?, ?, ?)""",
                    (table_name, file_path.name, sheet_name, len(headers), inserted, file_hash),
                )
                conn.commit()

                tables_created.append({"table": table_name, "columns": headers, "rows": inserted})
                total_rows += inserted
                logger.info("Exam DB: imported %d rows into '%s'", inserted, table_name)

        except Exception as e:
            logger.exception("Excel ingestion failed")
            conn.rollback()
            conn.close()
            return {"status": "error", "message": f"导入失败: {e}"}

        conn.close()
        return {
            "status": "ok",
            "tables_created": tables_created,
            "total_rows": total_rows,
            "message": f"成功导入 {len(tables_created)} 张表，共 {total_rows} 行",
        }

    # ===== NL → SQL → Results =====

    NL2SQL_PROMPT = """你是一个 SQL 查询生成器。根据以下数据库表结构和用户问题，生成一个 SQLite SELECT 查询语句。

## 数据库表结构

{tables_schema}

## 示例数据（每张表前2行）

{sample_data}

## 用户问题

{question}

## 规则
1. 只生成 SELECT 语句，不要任何其他内容
2. 列名用双引号包裹（如 "学校"）
3. 使用 LIKE 做模糊匹配（如 WHERE "学校" LIKE '%重邮%' OR "学校" LIKE '%重庆邮电大学%'）
4. 如果用户问"分数线"，查包含"复试线"或"分数线"的列
5. 如果用户问"报录比"，查包含"报录比"的列
6. 如果用户问"招生人数"或"录取人数"，查对应列
7. 如果用户做对比（"A和B哪个好考"），查两个学校的全部数据
8. 如果列的值看起来是数字但存为TEXT，用 CAST("列名" AS INTEGER) 做数值比较
9. 最多返回 {max_rows} 行
10. 如果问题与表中数据完全无关，回复 NO_MATCH

只输出 SQL 语句（或 NO_MATCH），不要 markdown 代码块，不要解释。"""

    def query(
        self,
        question: str,
        llm_fn=None,
    ) -> List[dict]:
        """自然语言查询 → SQL 生成 → 执行 → 返回结果。

        Args:
            question: 用户自然语言问题
            llm_fn: LLM 调用函数 (prompt, temperature, max_tokens) -> str

        Returns:
            List[dict]: 结构化结果，格式兼容 VectorStore.search() 的返回
        """
        # 0. 检查是否有可用表
        tables = self.get_tables()
        if not tables:
            logger.warning("Exam DB has no tables, cannot query")
            return []

        # 1. 构建 prompt：表结构 + 样本数据
        schema_text = self._build_schema_prompt(tables)
        sample_text = self._build_sample_prompt(tables)

        prompt = self.NL2SQL_PROMPT.format(
            tables_schema=schema_text,
            sample_data=sample_text,
            question=question,
            max_rows=MAX_RESULT_ROWS,
        )

        # 2. 调用 LLM 生成 SQL
        if llm_fn is None:
            logger.warning("No LLM function provided for NL2SQL, falling back")
            return []

        try:
            sql = llm_fn(prompt, temperature=0.0, max_tokens=512)
            sql = self._sanitize_sql(sql)
        except Exception as e:
            logger.exception("NL2SQL generation failed: %s", e)
            return []

        if sql == "NO_MATCH" or not sql:
            return []

        # 3. 安全检查
        if not self._validate_sql(sql):
            logger.warning("NL2SQL blocked unsafe SQL: %s", sql)
            return []

        # 4. 执行查询
        try:
            conn = self._get_conn()
            cursor = conn.execute(sql)
            rows = cursor.fetchall()
            conn.close()
        except Exception as e:
            logger.warning("SQL execution failed: %s | SQL: %s", e, sql)
            return []

        # 5. 格式化为标准结果
        return self._format_results(rows, sql)

    # ===== 工具方法 =====

    def get_tables(self) -> List[dict]:
        """获取所有数据表信息（排除元数据表）。"""
        conn = self._get_conn()
        tables = []
        try:
            cursor = conn.execute(
                "SELECT table_name, source_file, sheet_name, column_count, row_count "
                "FROM _exam_meta ORDER BY table_name"
            )
            for row in cursor:
                # 获取列信息
                col_cursor = conn.execute(f'PRAGMA table_info("{row["table_name"]}")')
                columns = [
                    {"name": c["name"], "type": c["type"]}
                    for c in col_cursor
                    if c["name"] != "id"
                ]
                tables.append({
                    "table_name": row["table_name"],
                    "source_file": row["source_file"],
                    "sheet_name": row["sheet_name"],
                    "columns": columns,
                    "row_count": row["row_count"],
                })
        except Exception as e:
            logger.error("Failed to get tables: %s", e)
        finally:
            conn.close()
        return tables

    def get_stats(self) -> dict:
        """获取数据库统计信息。"""
        tables = self.get_tables()
        return {
            "total_tables": len(tables),
            "total_rows": sum(t["row_count"] for t in tables),
            "tables": [
                {"name": t["table_name"], "rows": t["row_count"], "columns": len(t["columns"])}
                for t in tables
            ],
        }

    def clear(self):
        """清空所有考试数据表。"""
        conn = self._get_conn()
        tables = self.get_tables()
        for t in tables:
            conn.execute(f'DROP TABLE IF EXISTS "{t["table_name"]}"')
        conn.execute("DELETE FROM _exam_meta")
        conn.commit()
        conn.close()
        logger.info("Exam DB cleared: %d tables dropped", len(tables))

    # ===== 内部方法 =====

    def _build_schema_prompt(self, tables: List[dict]) -> str:
        """构建表结构的 prompt 文本。"""
        parts = []
        for t in tables:
            cols = ", ".join(f'"{c["name"]}" TEXT' for c in t["columns"])
            parts.append(f'表名: "{t["table_name"]}" ({t["row_count"]} 行)')
            parts.append(f"  列: {cols}")
        return "\n".join(parts)

    def _build_sample_prompt(self, tables: List[dict]) -> str:
        """构建示例数据的 prompt 文本。"""
        conn = self._get_conn()
        parts = []
        try:
            for t in tables:
                cols = [c["name"] for c in t["columns"]]
                if not cols:
                    continue
                col_str = ", ".join(f'"{c}"' for c in cols)
                cursor = conn.execute(
                    f'SELECT {col_str} FROM "{t["table_name"]}" LIMIT 2'
                )
                rows = cursor.fetchall()
                parts.append(f'--- "{t["table_name"]}" ---')
                for row in rows:
                    parts.append(" | ".join(str(row[c]) for c in cols if c in row.keys()))
        finally:
            conn.close()
        return "\n".join(parts)

    def _sanitize_sql(self, raw_sql: str) -> str:
        """清洗 LLM 输出：去掉 markdown 代码块、语言标签、前后空白、尾部分号。"""
        sql = raw_sql.strip()

        # 去掉 markdown 代码块（含语言标签如 ```sql）
        if "```" in sql:
            # 方法：找 ``` 之间的内容，跳过语言标签行
            parts = sql.split("```")
            for p in parts:
                p = p.strip()
                if not p:
                    continue
                # 跳过纯语言标签（sql, SQL, json 等）
                if p.lower() in ("sql", "json", "python", "text", ""):
                    continue
                # 如果以语言标签开头，去掉第一行
                if p.split("\n")[0].strip().lower() in ("sql", "json", "python"):
                    p = "\n".join(p.split("\n")[1:]).strip()
                if p and (p.upper().startswith("SELECT") or p == "NO_MATCH"):
                    sql = p
                    break

        # 去掉前缀废话（如 "这是查询："）
        select_pos = sql.upper().find("SELECT")
        if select_pos > 0:
            sql = sql[select_pos:]

        # 去掉尾部多余内容：分号、反引号、空白
        sql = sql.rstrip(";").strip()
        sql = sql.rstrip("`").strip()

        return sql

    def _validate_sql(self, sql: str) -> bool:
        """SQL 安全校验：只允许 SELECT 语句。"""
        sql_upper = sql.upper().replace("\n", " ").replace("\t", " ")
        if not sql_upper.startswith("SELECT"):
            return False
        for keyword in FORBIDDEN_SQL_KEYWORDS:
            # 用词边界检测（避免误杀含 keyword 的合法列名）
            if re.search(r'\b' + re.escape(keyword) + r'\b', sql_upper):
                return False
        return True

    def _format_results(self, rows: List[sqlite3.Row], sql: str) -> List[dict]:
        """将 SQL 查询结果格式化为与 VectorStore.search() 兼容的结构。"""
        results = []
        for row in rows:
            # 将 Row 对象转为 dict
            row_dict = {k: row[k] for k in row.keys()}
            # 构建可读文本
            content_parts = [f"{k}: {v}" for k, v in row_dict.items() if v and k != "id"]
            content = "，".join(content_parts)

            results.append({
                "content": content,
                "metadata": {
                    "source": "exam_database",
                    "type": "nl2sql_result",
                    "record": json.dumps(row_dict, ensure_ascii=False),
                    "sql": sql,
                },
                "score": 1.0,
                "collection": "exam",
            })

        return results[:MAX_RESULT_ROWS]

    def _sanitize_table_name(self, name: str) -> str:
        """清洗表名：只保留中文、字母、数字、下划线。"""
        cleaned = re.sub(r'[^\w一-鿿]', '_', name)
        return cleaned.strip('_') or "exam_data"

    def _sanitize_column_name(self, name: str) -> str:
        """清洗列名。"""
        return name.strip().replace('"', '').replace("'", "") or "unnamed"

    @staticmethod
    def _file_hash(file_path: Path) -> str:
        import hashlib
        hasher = hashlib.md5()
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(8192), b""):
                hasher.update(chunk)
        return hasher.hexdigest()

    def _get_existing_hash(self, filename: str) -> Optional[str]:
        """获取已导入文件的哈希。"""
        conn = self._get_conn()
        cursor = conn.execute(
            "SELECT DISTINCT file_hash FROM _exam_meta WHERE source_file = ?",
            (filename,),
        )
        row = cursor.fetchone()
        conn.close()
        return row["file_hash"] if row else None


# ===== 便捷函数 =====

_exam_db: Optional[ExamDatabase] = None


def get_exam_db() -> ExamDatabase:
    """获取全局 ExamDatabase 单例。"""
    global _exam_db
    if _exam_db is None:
        _exam_db = ExamDatabase()
    return _exam_db


if __name__ == "__main__":
    from logger_config import get_logger
    get_logger(__name__)

    db = ExamDatabase()
    print(f"Tables: {db.get_stats()}")

    # 导入测试
    test_file = config.DOCUMENTS_DIR / "考研分数线及报录比数据.xlsx"
    if test_file.exists():
        result = db.ingest_excel(test_file, force=True)
        print(f"Ingest: {json.dumps(result, ensure_ascii=False, indent=2)}")
        print(f"After ingest: {db.get_stats()}")
